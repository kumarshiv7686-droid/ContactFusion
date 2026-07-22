#!/usr/bin/env python3
"""
consolidate_contacts.py
========================
Simple, fast, standalone tool. No database, no GUI framework, no wide
master schema -- just: scan a folder, pull out Name / Number / Address
from every file, dedupe by phone number, write clean output.

USAGE
-----
    python consolidate_contacts.py <source_folder> [output_folder]

If you don't pass arguments, it will ask you for the folder(s) when run.

WHAT IT DOES
------------
1. Recursively scans the source folder for .xlsx / .xlsm / .xls / .xlsb / .csv
2. For each file, figures out which column is Name, which is Address, and
   which is the phone Number:
     - Number: looks for a header containing mobile/phone/contact/cell/
       whatsapp/mob/tel. If NO such header exists, it falls back to
       scanning the actual cell values -- any column where most values are
       10-13 digit numbers (after stripping spaces/dashes/+) is treated as
       the Number column.
     - Name / Address: matched by header keywords only.
3. Streams rows (never loads a whole file into memory at once).
4. Keeps only rows that have a valid number (10-13 digits). Rows with no
   usable number are dropped, since the goal is a clean deduplicated
   number list.
5. Merges duplicate numbers (different formatting like "+91 98765 43210"
   vs "9876543210" vs "9876543210.0" are recognized as the same number).
   If one occurrence has a blank Name/Address and another has it filled
   in, the filled-in value is kept.
6. Writes output as .xlsx. If the unique row count exceeds Excel's
   1,048,576-row-per-sheet limit, it automatically splits into
   consolidated_part1.xlsx, consolidated_part2.xlsx, etc.

REQUIREMENTS
------------
    pip install openpyxl
Optional (only needed if you actually have these file types):
    pip install xlrd      # for legacy .xls
    pip install pyxlsb    # for .xlsb
"""
import csv
import os
import re
import sys
import time
from core import progress
try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install it with:  pip install openpyxl")
    sys.exit(1)

# --------------------------------------------------------------------------
# Config
# --------------------------------------------------------------------------
EXCEL_MAX_ROWS = 1_048_576
ROWS_PER_OUTPUT_FILE = EXCEL_MAX_ROWS - 1  # leave room for the header row

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xls", ".xlsb", ".csv"}

NAME_KEYWORDS = ["name"]
NAME_EXCLUDE = ["father", "mother", "husband", "company", "firm", "bank", "product", "scheme"]

ADDRESS_KEYWORDS = ["address", "addr", "location", "residence"]

NUMBER_KEYWORDS = ["mobile", "phone", "contact", "cell", "whatsapp", "mob", "tel"]

# How many data rows to sample when guessing the Number column by content
CONTENT_SCAN_SAMPLE_SIZE = 30
# Fraction of sampled values that must look like a 10-13 digit number
CONTENT_SCAN_THRESHOLD = 0.5


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------
def norm_header(s) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s).lower()) if s is not None else ""


def digits_only(value) -> str:
    """Strip everything except digits from a raw cell value. Also removes
    a trailing '.0' that Excel adds when a number is stored as a float."""
    if value is None:
        return ""
    s = str(value).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return re.sub(r"\D", "", s)


def is_number_like(value) -> bool:
    d = digits_only(value)
    return 10 <= len(d) <= 13


def dedup_key(value) -> str:
    """Normalize a valid number down to its last 10 digits, so formatting
    and country-code differences (+91, 0-prefix, spaces) don't create
    false 'different' numbers."""
    d = digits_only(value)
    return d[-10:] if len(d) >= 10 else d


def display_number(value) -> str:
    """Same normalization as dedup_key -- always show a clean number
    (country code / extra prefix digits stripped) rather than whatever
    raw digit count happened to be in the source cell."""
    return dedup_key(value)


def pick_header_column(headers, keywords, exclude=None):
    exclude = exclude or []
    normed = [norm_header(h) for h in headers]
    for idx, n in enumerate(normed):
        if any(k in n for k in keywords) and not any(x in n for x in exclude):
            return idx
    return None


def guess_number_column_by_content(sample_rows, num_columns):
    """Given a sample of data rows, find the column index most likely to
    be a phone number column based on cell content alone."""
    if not sample_rows or num_columns == 0:
        return None
    best_idx, best_ratio = None, 0.0
    for col in range(num_columns):
        values = [row[col] for row in sample_rows if col < len(row)]
        non_blank = [v for v in values if v not in (None, "")]
        if not non_blank:
            continue
        hits = sum(1 for v in non_blank if is_number_like(v))
        ratio = hits / len(non_blank)
        if ratio > best_ratio:
            best_ratio = ratio
            best_idx = col
    return best_idx if best_ratio >= CONTENT_SCAN_THRESHOLD else None


def find_header_row(rows_preview):
    """First row with >=2 non-blank cells is treated as the header row."""
    for idx, row in enumerate(rows_preview):
        non_blank = [c for c in row if c not in (None, "")]
        if len(non_blank) >= 2:
            return idx
    for idx, row in enumerate(rows_preview):
        if any(c not in (None, "") for c in row):
            return idx
    return 0


# --------------------------------------------------------------------------
# File readers -- each yields (headers, row) tuples, streaming, one file
# at a time, one sheet at a time.
# --------------------------------------------------------------------------
def read_xlsx_rows(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            row_iter = ws.iter_rows(values_only=True)
            preview = []
            header = None
            for i, row in enumerate(row_iter):
                if header is None:
                    preview.append(list(row))
                    if i < 4:
                        continue
                    hidx = find_header_row(preview)
                    header = [str(c).strip() if c is not None else "" for c in preview[hidx]]
                    for leftover in preview[hidx + 1:]:
                        yield sheet_name, header, leftover
                    continue
                yield sheet_name, header, list(row)
            if header is None and preview:
                hidx = find_header_row(preview)
                header = [str(c).strip() if c is not None else "" for c in preview[hidx]]
                for leftover in preview[hidx + 1:]:
                    yield sheet_name, header, leftover
    finally:
        wb.close()


def read_csv_rows(path):
    def _read(encoding):
        with open(path, "r", newline="", encoding=encoding, errors="replace") as f:
            reader = csv.reader(f)
            preview = []
            header = None
            for i, row in enumerate(reader):
                if header is None:
                    preview.append(row)
                    if i < 4:
                        continue
                    hidx = find_header_row(preview)
                    header = [c.strip() for c in preview[hidx]]
                    for leftover in preview[hidx + 1:]:
                        yield "<csv>", header, leftover
                    continue
                yield "<csv>", header, row
            if header is None and preview:
                hidx = find_header_row(preview)
                header = [c.strip() for c in preview[hidx]]
                for leftover in preview[hidx + 1:]:
                    yield "<csv>", header, leftover

    try:
        yield from _read("utf-8-sig")
    except UnicodeDecodeError:
        yield from _read("latin-1")


def read_xls_rows(path):
    try:
        import xlrd
    except ImportError:
        print(f"  SKIPPED (install 'xlrd' to read .xls files): {path}")
        return
    wb = xlrd.open_workbook(path)
    for sheet in wb.sheets():
        rows = [sheet.row_values(r) for r in range(sheet.nrows)]
        if not rows:
            continue
        hidx = find_header_row(rows)
        header = [str(c).strip() for c in rows[hidx]]
        for row in rows[hidx + 1:]:
            yield sheet.name, header, row


def read_xlsb_rows(path):
    try:
        from pyxlsb import open_workbook
    except ImportError:
        print(f"  SKIPPED (install 'pyxlsb' to read .xlsb files): {path}")
        return
    with open_workbook(path) as wb:
        for sheet_name in wb.sheets:
            with wb.get_sheet(sheet_name) as sheet:
                rows = [[c.v for c in row] for row in sheet.rows()]
                if not rows:
                    continue
                hidx = find_header_row(rows)
                header = [str(c).strip() if c is not None else "" for c in rows[hidx]]
                for row in rows[hidx + 1:]:
                    yield sheet_name, header, row


def read_any(path, ext):
    if ext in (".xlsx", ".xlsm"):
        yield from read_xlsx_rows(path)
    elif ext == ".csv":
        yield from read_csv_rows(path)
    elif ext == ".xls":
        yield from read_xls_rows(path)
    elif ext == ".xlsb":
        yield from read_xlsb_rows(path)


# --------------------------------------------------------------------------
# Main pipeline
# --------------------------------------------------------------------------
def find_files(root_folder):
    for dirpath, _dirnames, filenames in os.walk(root_folder):
        for fname in filenames:
            ext = os.path.splitext(fname)[1].lower()
            if ext in SUPPORTED_EXTENSIONS:
                yield os.path.join(dirpath, fname), ext


def unique_output_path(directory, filename):
    """If `filename` already exists (e.g. still open in Excel, or left
    over from a previous run), find the next available name by appending
    _2, _3, _4, etc. before the extension."""
    stem, ext = os.path.splitext(filename)
    candidate = os.path.join(directory, filename)
    n = 2
    while os.path.exists(candidate):
        candidate = os.path.join(directory, f"{stem}_{n}{ext}")
        n += 1
    return candidate


def save_with_retry(wb, directory, filename, max_attempts=30):
    """Save the workbook, automatically trying filename_2, filename_3, ...
    both when the target name already exists AND when saving fails with a
    PermissionError (typically because the existing file is open in Excel
    or locked by another program)."""
    out_path = unique_output_path(directory, filename)
    attempt = 0
    while True:
        try:
            wb.save(out_path)
            return out_path
        except PermissionError:
            attempt += 1
            if attempt >= max_attempts:
                raise
            stem, ext = os.path.splitext(filename)
            out_path = os.path.join(directory, f"{stem}_{attempt + 1}{ext}")
            print(f"    '{filename}' is locked or in use -- trying '{os.path.basename(out_path)}' instead")


def consolidate(source_folder, output_folder):
    os.makedirs(output_folder, exist_ok=True)

    files = list(find_files(source_folder))
    progress.start(len(files))
    print(f"Found {len(files)} file(s) to process.\n")

    contacts = {}  # dedup_key -> [name, number_display, address]
    total_rows_seen = 0
    total_rows_kept = 0
    start = time.time()

    for file_idx, (path, ext) in enumerate(files, start=1):
        progress.update_file(os.path.basename(path))
        fname = os.path.basename(path)
        print(f"[{file_idx}/{len(files)}] {fname}")

        file_valid_rows = 0
        current_sheet_key = None
        name_idx = addr_idx = num_idx = None
        pending_rows_before_detection = []
        detected_for_sheet = False

        try:
            for sheet_name, header, row in read_any(path, ext):
                total_rows_seen += 1
                sheet_key = (sheet_name, tuple(header))

                if sheet_key != current_sheet_key:
                    # New sheet encountered -- headers may be completely
                    # different from the previous sheet. First, flush any
                    # rows still buffered from the PREVIOUS sheet (it ended
                    # before hitting the content-detection sample size).
                    if current_sheet_key is not None and not detected_for_sheet and pending_rows_before_detection:
                        num_idx = guess_number_column_by_content(
                            pending_rows_before_detection,
                            max((len(r) for r in pending_rows_before_detection), default=0))
                        note = " <- VERIFY this isn't Aadhaar/PAN/GST/an ID column" if num_idx is not None else ""
                        print(f"    [{current_sheet_key[0]}] Number="
                              f"{('content-detected col ' + str(num_idx) + note) if num_idx is not None else '(NOT FOUND)'}")
                        for buffered_row in pending_rows_before_detection:
                            kept = _process_row(buffered_row, name_idx, num_idx, addr_idx, contacts)
                            total_rows_kept += kept
                            file_valid_rows += kept

                    # Start fresh detection for the new sheet
                    current_sheet_key = sheet_key
                    name_idx = pick_header_column(header, NAME_KEYWORDS, NAME_EXCLUDE)
                    addr_idx = pick_header_column(header, ADDRESS_KEYWORDS)
                    num_idx = pick_header_column(header, NUMBER_KEYWORDS)
                    pending_rows_before_detection = []
                    detected_for_sheet = num_idx is not None
                    if detected_for_sheet:
                        print(f"    [{sheet_name}] Name={header[name_idx] if name_idx is not None else '(none)'} | "
                              f"Number={header[num_idx]} | "
                              f"Address={header[addr_idx] if addr_idx is not None else '(none)'}")

                if not detected_for_sheet:
                    pending_rows_before_detection.append(row)
                    if len(pending_rows_before_detection) < CONTENT_SCAN_SAMPLE_SIZE:
                        continue
                    num_idx = guess_number_column_by_content(pending_rows_before_detection, len(header))
                    detected_for_sheet = True
                    note = " <- VERIFY this isn't Aadhaar/PAN/GST/an ID column" if num_idx is not None else ""
                    print(f"    [{sheet_name}] Name={header[name_idx] if name_idx is not None else '(none)'} | "
                          f"Number={('content-detected col ' + str(num_idx) + note) if num_idx is not None else '(NOT FOUND)'} | "
                          f"Address={header[addr_idx] if addr_idx is not None else '(none)'}")
                    for buffered_row in pending_rows_before_detection:
                        kept = _process_row(buffered_row, name_idx, num_idx, addr_idx, contacts)
                        total_rows_kept += kept
                        file_valid_rows += kept
                    continue

                kept = _process_row(row, name_idx, num_idx, addr_idx, contacts)
                total_rows_kept += kept
                file_valid_rows += kept

            # File ended -- flush any rows still buffered from the LAST sheet
            if current_sheet_key is not None and not detected_for_sheet and pending_rows_before_detection:
                num_idx = guess_number_column_by_content(
                    pending_rows_before_detection,
                    max((len(r) for r in pending_rows_before_detection), default=0))
                note = " <- VERIFY this isn't Aadhaar/PAN/GST/an ID column" if num_idx is not None else ""
                print(f"    [{current_sheet_key[0]}] Number="
                      f"{('content-detected col ' + str(num_idx) + note) if num_idx is not None else '(NOT FOUND)'}")
                for buffered_row in pending_rows_before_detection:
                    kept = _process_row(buffered_row, name_idx, num_idx, addr_idx, contacts)
                    total_rows_kept += kept
                    file_valid_rows += kept

            print(f"    -> {file_valid_rows:,} valid-number rows from this file")
            progress.update_processed(
                rows=file_valid_rows,
                unique=len(contacts),
                duplicates=total_rows_kept - len(contacts)
            )
            progress.file_completed()    
        except Exception as exc:  # noqa: BLE001
            print(f"    ERROR reading {fname}: {exc}")
            continue

    elapsed = time.time() - start
    print(f"\nScanned {total_rows_seen:,} rows, kept {total_rows_kept:,} rows with a valid number.")
    print(f"Unique numbers: {len(contacts):,}  (took {elapsed:.1f}s)\n")

    # Write output, splitting at Excel's row limit
    all_rows = list(contacts.values())
    total = len(all_rows)
    multi_part = total > ROWS_PER_OUTPUT_FILE
    idx, part, outputs = 0, 1, []
    while idx < total or part == 1:
        chunk = all_rows[idx: idx + ROWS_PER_OUTPUT_FILE]
        filename = f"consolidated_part{part}.xlsx" if multi_part else "consolidated.xlsx"
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet("Contacts")
        ws.append(["Name", "Number", "Address"])
        for row in chunk:
            ws.append(row)
        out_path = save_with_retry(wb, output_folder, filename)
        outputs.append(out_path)
        print(f"Wrote {len(chunk):,} rows -> {out_path}")
        idx += ROWS_PER_OUTPUT_FILE
        part += 1
        if total == 0:
            break

    if outputs:
        progress.complete(outputs)
    return outputs

def _process_row(row, name_idx, num_idx, addr_idx, contacts) -> int:
    """Extract Name/Number/Address from one raw row and merge it into
    `contacts` if it has a valid number. Returns 1 if the row was kept."""
    if num_idx is None:
        return 0
    num_val = row[num_idx] if num_idx < len(row) else None
    if not is_number_like(num_val):
        return 0

    key = dedup_key(num_val)
    name_val = row[name_idx] if (name_idx is not None and name_idx < len(row)) else None
    addr_val = row[addr_idx] if (addr_idx is not None and addr_idx < len(row)) else None
    name_str = "" if name_val is None else str(name_val).strip()
    addr_str = "" if addr_val is None else str(addr_val).strip()
    number_str = display_number(num_val)

    if key not in contacts:
        contacts[key] = [name_str, number_str, addr_str]
    else:
        existing = contacts[key]
        if not existing[0] and name_str:
            existing[0] = name_str
        if not existing[2] and addr_str:
            existing[2] = addr_str
    return 1


# --------------------------------------------------------------------------
if __name__ == "__main__":
    if len(sys.argv) >= 2:
        src = sys.argv[1]
    else:
        src = input("Source folder to scan: ").strip().strip('"')

    if len(sys.argv) >= 3:
        out = sys.argv[2]
    else:
        out = input("Output folder (Enter for './consolidated_output'): ").strip().strip('"')
        if not out:
            out = os.path.join(os.getcwd(), "consolidated_output")

    if not os.path.isdir(src):
        print(f"ERROR: source folder not found: {src}")
        sys.exit(1)

    consolidate(src, out)
    print("\nDone.")